from flask import Flask, request, jsonify, render_template, session
from openpyxl import Workbook, load_workbook
import os
import uuid

app = Flask(__name__)
app.secret_key = 'super_secret_key'

# Имя файла Excel, куда будут сохраняться ответы
EXCEL_FILE = r'C:\Codes\IMBA_DP.xlsx'

# Если файл Excel не существует, создаем его с нужными колонками
if not os.path.exists(EXCEL_FILE):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(['User ID', 'Answer', 'Variable', 'Answer D', 'Completed', 'DP'])  # Добавляем DP в заголовки
    workbook.save(EXCEL_FILE)

def generate_user_id():
    """Генерирует уникальный UUID для пользователя."""
    return str(uuid.uuid4())

def get_user_row(user_id, sheet):
    """Ищет строку пользователя по user_id в Excel и возвращает её индекс."""
    for row in sheet.iter_rows(min_row=2, values_only=False):
        if row[0].value == user_id:
            return row[0].row  # Возвращаем индекс строки
    return None  # Если user_id не найден

@app.route('/')
def index():
    if 'user_id' not in session:
        session['user_id'] = generate_user_id()
    return render_template('index.html')

@app.route('/survey')
def survey():
    return render_template('survey.html')

@app.route('/submit', methods=['POST'])
def submit():
    data = request.get_json()
    answer = data.get('answer')
    user_id = session.get('user_id')

    if answer and user_id:
        workbook = load_workbook(EXCEL_FILE)
        sheet = workbook.active

        user_row = get_user_row(user_id, sheet)
        
        if user_row:
            sheet[f'B{user_row}'] = answer
            if 'подопрос 2' in answer.lower():
                sheet[f'D{user_row}'] = answer
                sheet[f'E{user_row}'] = 1
            
            # Увеличиваем DP
            dp_value = sheet[f'F{user_row}'].value or 0
            dp_value += 10
            # Проверяем, выполнены ли 2 подопроса
            completed_count = sheet[f'E{user_row}'].value or 0
            if completed_count >= 2:
                dp_value += 20
            sheet[f'F{user_row}'] = dp_value

        else:
            # Инициализация новой строки
            dp_value = 10  # Начальное значение DP
            completed = 1 if 'подопрос 2' in answer.lower() else 0
            sheet.append([user_id, answer, 0, answer if 'подопрос 2' in answer.lower() else '', completed, dp_value])  # Инициализация

        workbook.save(EXCEL_FILE)
        workbook.close()
        return jsonify({'status': 'success'}), 200
    else:
        return jsonify({'status': 'error', 'message': 'No answer or user ID provided'}), 400

@app.route('/update_variable', methods=['POST'])
def update_variable():
    user_id = session.get('user_id')
    if not user_id:
        return jsonify({'status': 'error', 'message': 'User ID not found in session'}), 400

    workbook = load_workbook(EXCEL_FILE)
    sheet = workbook.active
    user_row = get_user_row(user_id, sheet)

    if user_row:
        # Увеличиваем значение переменной
        current_value = sheet[f'C{user_row}'].value or 0
        sheet[f'C{user_row}'] = current_value + 1

        # Увеличиваем DP на 10
        dp_value = sheet[f'F{user_row}'].value or 0
        dp_value += 10

        # Проверка выполнения условий обоих подопросов
        completed_count = sheet[f'E{user_row}'].value or 0
        if completed_count >= 2:
            dp_value += 20  # Дополнительное увеличение DP

        # Сохраняем обновленное значение DP
        sheet[f'F{user_row}'] = dp_value

        workbook.save(EXCEL_FILE)
        workbook.close()
        return jsonify({
            'status': 'success',
            'message': f'Variable updated to {current_value + 1}, DP updated to {dp_value}',
            'variable': current_value + 1,
            'DP': dp_value
        }), 200
    else:
        workbook.close()
        return jsonify({'status': 'error', 'message': 'User ID not found in database'}), 404


@app.route('/check_variable', methods=['GET'])
def check_variable():
    user_id = session.get('user_id')
    if not user_id:
        return jsonify({'status': 'error', 'message': 'User ID not found in session'}), 400

    workbook = load_workbook(EXCEL_FILE, data_only=True)
    sheet = workbook.active
    user_row = get_user_row(user_id, sheet)

    if user_row:
        variable_value = sheet[f'C{user_row}'].value
        dp_value = sheet[f'F{user_row}'].value
        workbook.close()
        return jsonify({'variable': variable_value, 'DP': dp_value}), 200
    else:
        workbook.close()
        return jsonify({'status': 'error', 'message': 'User ID not found in database'}), 404

@app.route('/withdraw')
def withdraw():
    return render_template('withdraw.html')

@app.route('/withdraw_dp', methods=['POST'])
def withdraw_dp():
    data = request.get_json()
    dp_requested = int(data.get('dp'))
    user_id = session.get('user_id')

    if not user_id:
        return jsonify({'status': 'error', 'message': 'User ID not found in session'}), 400

    workbook = load_workbook(EXCEL_FILE)
    sheet = workbook.active
    user_row = get_user_row(user_id, sheet)

    if user_row:
        current_dp = sheet[f'F{user_row}'].value or 0
        if dp_requested > current_dp:
            workbook.close()
            return jsonify({'status': 'error', 'message': 'Слишком много DP ввёл, не гойда'}), 400
        else:
            # Обновляем DP в базе данных
            sheet[f'F{user_row}'] = current_dp - dp_requested
            workbook.save(EXCEL_FILE)
            workbook.close()
            return jsonify({'status': 'success', 'message': f'Вы успешно вывели {dp_requested} DP, ГОЙДА'}), 200
    else:
        workbook.close()
        return jsonify({'status': 'error', 'message': 'User ID not found in database'}), 404

if __name__ == '__main__':
    app.run(debug=True)
